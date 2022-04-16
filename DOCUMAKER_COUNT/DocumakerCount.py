from multiprocessing import allow_connection_pickling, freeze_support
import configparser, logging
import shutil, os,os.path,fnmatch, re
from io import StringIO
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
import pandas as pd
import concurrent
from concurrent.futures import ProcessPoolExecutor
import ibm_db
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import date, datetime
import warnings,sys


warnings.filterwarnings("ignore")

# SETTUP FOR LOGGING
logger = logging.getLogger('DocumakerCount')    
logger.setLevel(logging.INFO)
formater = logging.Formatter('%(asctime)s:%(levelname)s:%(name)s:%(message)s')
fileHandler = logging.FileHandler('LOGS\\DocumakerCount.log')
fileHandler.setFormatter(formater)
logger.addHandler(fileHandler)

# SETUP CONFIG PARSER TO READ VALUES FROM INI FILE
config = configparser.RawConfigParser()
config.read('countproperties.ini')

pdflocation = f'downloadpdf\\'
resultpath = f'Results\\'

hostName = config['DBCONNECT']['host']
port = config['DBCONNECT']['port']
dbName = config['DBCONNECT']['dbname']
uDname = config['DBCONNECT']['dbuser']
passWord = config['DBCONNECT']['dbpass']

# MOVE OLD FILE FROM RESULT FOLDER TO BACKUP FOLDER
def move_file():
    backup = f'Results\Backup'
    file = fnmatch.filter(os.listdir(resultpath), '*.xlsx')
    if len(file) > 0:
        shutil.move(f'{resultpath}{file[0]}',backup)
        logger.info('Line No: 48: Program Started. Excel File moved to Backup Folder.')

def read_pdf(filename,Startpagevalue,Endrangevalue):      
    output_string01 = StringIO()      
    
    with open(f'{filename}', 'rb') as in_file:
        in_file = open(f'{filename}', 'rb')
        parser = PDFParser(in_file)        
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()        
        device = TextConverter(rsrcmgr, output_string01, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        

        for i, page in enumerate(PDFPage.create_pages(doc)):
            if Startpagevalue=='0' and Endrangevalue == '0':
                read_position = output_string01.tell()
                interpreter.process_page(page)
                output_string01.seek(read_position, 0)
                page_text = output_string01.read()
                eachpagelist = re.split('\n+', page_text)
                eachpagelist = [x for x in eachpagelist if int(len(x)) > 1]                
                yield eachpagelist
            else:
                if i>=int(Startpagevalue) and i<=int(Endrangevalue):
                    read_position = output_string01.tell()
                    interpreter.process_page(page)
                    output_string01.seek(read_position, 0)
                    page_text = output_string01.read()
                    eachpagelist = re.split('\n+', page_text)
                    eachpagelist = [x for x in eachpagelist if int(len(x)) > 1]                    
                    yield eachpagelist
                else:
                    pass
    

def extract_contents(filename,Startpagevalue,Endrangevalue):
    path = f'downloadpdf\\'
    print(f'Processing PDF: {filename}')
    datas = read_pdf(f'{path}{filename}',Startpagevalue,Endrangevalue)
    # print(len(list(datas)))
    

    letternameandvalue = {'notice third party offer':['important information about your policy - please review','notice third party offer'],'confirmation premium payment':['premium payment confirmation'],
    'confirmation loan collateral interest credit':['loan collateral interest credit confirmation'],'confirmation loan interest capitalization':['loan interest capitalization confirmation'],'bill loan interest only':['notice of loan interest payment due','face amount'],'confirmation apl':['automatic premium loan confirmation'],'notice net gain dividends':['important tax information – immediate attention needed'],'confirmation discontinue accelerated payment-premium offset':["discontinue accelerated payment confirmation"],'notice term expiration':['notice term expiration'],'quote withdrawl':['policy withdrawal quote'],'audit':['your request has been received – please review','audit'],'1035 exchange confirmation':['1035 exchange confirmation'],'acknowledgement of life insurance letter':['statement of coverage – please review'],'accelerated payment confirmation':['accelerated payment confirmation','under this payment arrangement, a withdrawal from your policy’s'],'accelerated payment quote':['accelerated payment quote'],'address change confirmation':['address change confirmation'],'annual policy statement':['annual policy statement'],'auto reinstatement':['notice of policy lapse','thank you for your premium payment'],'automatic premium loan confirmation':['automatic premium loan confirmation'],'bill grace prelapse':['notice of overdue payment','refer to your original policy for specific'],
    'ca beneficiary notification':['ca beneficiary notification','important information about your policy– please review'],'ca consumer letter':['important insurance disclosure – please review','ca consumer letter'],'confirmation frequency change':['payment change confirmation','the frequency of'],'confirmation reinstatement':['reinstatement confirmation'],'dividend option change':['dividend transaction confirmation'],'dividend option change t':['dividend transfer confirmation'],'dividend option change o':['dividend option change confirmation'],'expiration of rider-benefit r':['your policy contains a rider – please review'],'electronic payment confirmation':['electronic payment confirmation'],'explanation of benefits':['explanation of benefits','total control account (tca)'],'fl consumer letter':['fl consumer letter','important insurance disclosure – please review'],'fully paid-up confirmation':['fully paid-up confirmation'],'life quote nonforfeiture':['reduced paid-up insurance quote','life quote nonforfeiture'],'loan exceeds cash value termination':['important information – your policy has been cancelled'],'lapse notification':['important notification – your policy has lapsed'],'loan interest change confirmation':['loan interest change confirmation'],'loan repayment confirmation':['loan repayment confirmation'],'maturity notification':['your reply is important – please review and respond'],'nh consumer letter':['nh consumer letter','Important insurance disclosure – Please review'],'notice impact of loan or withdrawal on accelerated payment':['important information – please review'],'notice child':['important information about your policy– please review','this notice is to remind you that you have a child term rider'],'notice nsf':['your payment has been returned– immediate attention required'],'notice of minimum loan repayment':['notice of minimum loan repayment'],'owner and beneficiary change':['policy update confirmation'],'owner and beneficiary change c':['collateral assignment confirmation'],'owner and beneficiary change n':['name change confirmation'],'owner and beneficiary change o':['owner confirmation'],'owner and beneficiary change b':['beneficiary change confirmation'],'policy cash value quote':['policy cash value quote'],'policy dividend history' :['policy dividend history'],'policy loan confirmation':['policy loan confirmation'],'policy loan history':['policy loan history'],
    'policy loan quote':['policy loan quote'],'policy surrender confirmation':['policy surrender confirmation'],'premium payment history':['premium payment history'],'quote term policy':['quote term policy','term policy value quote'],'ri consumer letter':['ri consumer letter','important insurance disclosure – please review'],'recorded nonforfeiture confirmation':['recorded nonforfeiture confirmation'],'skip recovery privacy notice':['privacy policy – please review'],'stale check':['your request has been received – please review','stale dated checks'],'tax identification confirmation':['tax identification number update confirmation'],'third party designation confirmation':['third party designation confirmation'],'wi consumer letter':['important insurance disclosure – please review','wi consumer letter'],'wma check':['pay to the order of','reason:'],'expiration of rider-benefit':['your policy contains a benefit – please review'],
    'notice premium or rider premium change':['your premium is changing – please review'],'suspend billing confirmation':['suspend billing confirmation'],'suspend billing confirmation r':['resume billing confirmation'],'notice of payment due':['notice of payment due','face amount'],'confirmation planned premium change':['payment change confirmation','we processed your premium change']}

    # letternameandvalue = {'fl consumer letter':['fl consumer letter','important insurance disclosure – please review']
    # }
    
    letterscount = 0
    all_data = []
    for data in datas:
        listToStrvalue = ' '.join([str(elem) for elem in data])
        
        # Remove unwanted spaces
        listToStrvalue = re.sub('\s{2,}', ' ', listToStrvalue)
        # print(f'{listToStrvalue}')

        for key in letternameandvalue:
            letterheadinglist = letternameandvalue[key]
            lowerlistToStrvalue = listToStrvalue.lower()
            
            present = all(item in lowerlistToStrvalue for item in letterheadinglist)
            try:
                if present == True:
                    letterheading = letterheadinglist[0]
                    lettername = key
                    if(lettername in("stale check")):
                        if filename.__contains__("STALECHECK90"):
                            lettername = 'stalecheck90'
                        else:
                            lettername = 'stalecheck150'

                    letterscount = letterscount+1

                    if(lettername in {"bill loan interest only","bill loan interest only","notice of payment due", "bill grace prelapse a" , "bill grace prelapse" , 'notice of minimum loan repayment a' , 'notice of minimum loan repayment'}):
                        firstvalue = "policy number"

                    elif(lettername in {"lapse notification","lapse notificationa","confirmation discontinue accelerated payment-premium offset","dividend option change t","dividend option change o","dividend option change","recorded nonforfeiture confirmation","confirmation reinstatement","stalecheck90","stalecheck150","notice term expiration","notice nsf","notice impact of loan or withdrawal on accelerated payment","notice child","loan interest change confirmation","loan interest change confirmation a","fully paid-up confirmation","fully paid-up confirmation","auto reinstatement","acknowledgement of life insurance letter","loan exceeds cash value termination","skip recovery privacy notice","owner and beneficiary change","owner and beneficiary change c","owner and beneficiary change n","owner and beneficiary change o","owner and beneficiary change b","owner and beneficiary change a","owner and beneficiary change ca","owner and beneficiary change na","owner and beneficiary change oa","owner and beneficiary change ba","policy update confirmation","confirmation frequency change","ca beneficiary notification","loan exceeds cash value termination","expiration of rider-benefit","maturity notification","notice third party offer","notice premium or rider premium change","nh consumer letter","wi consumer letter","ri consumer letter","fl consumer letter","ca consumer letter"}):
                        if lowerlistToStrvalue.__contains__('policy:'):
                            firstvalue = "policy:"
                        else:
                            firstvalue = "company policy"

                    else:
                        firstvalue = 'policy number'
                    policynumber = " "


                    firstvalueindex =  lowerlistToStrvalue.index(firstvalue)+len(firstvalue)
                    secondvalueofindex = firstvalueindex+15
                    policystring = lowerlistToStrvalue[firstvalueindex:secondvalueofindex]
                    policystring = policystring.replace(":","").strip()
                    
                    res = " " in policystring
                    if(res == True):
                        policsplitlist = policystring.split()
                        policynumberc = policsplitlist[0]
                    else:
                        policynumberc = policystring
                    policynumber = policynumberc.upper()

                    # print(f'Policy Number: {policynumber}\n')
                    # logger.info(f'Policy Number: {policynumber}')
                    logger.info(f'Letter_Name: {lettername}, Policy_Number: {policynumber}')

                    row_dict = {
                        'Letter_Name' : lettername,
                        'Letter_Heading' : letterheading,
                        'Policy_No' : policynumber,
                        'Pdf_Name' : filename
                    }
                    all_data.append(row_dict)
                    
            
            except Exception as e:
                logging.error(f'Error while processing {filename} and {policynumber} because of:  {e}')
                continue
                  
    return all_data


def parallel_execution(params_list):    
    required_content = extract_contents(params_list[0],params_list[1],params_list[2])    
    return required_content


# METHOD TO RUN DB QUERY 
def runQuery(dbIBMConn,query):
    try:        
        stmtprp = ibm_db.exec_immediate(dbIBMConn, query)        
        rowData = ibm_db.fetch_assoc(stmtprp)       
 
    except Exception as e:
        print(f'ERROR runQueryIBM - {e}')
        logger.error(f'ERROR runQueryIBM - {e}')
        
    else:     
        logger.info('Query ran sucessfully. Tabulating the data...')     
        ibmResultList = []
        while rowData != False:
            ibmResultList.append(rowData)
            rowData = ibm_db.fetch_assoc(stmtprp)    
        return ibmResultList


def handle_db_query(dbconn,policyno): 
    # print(policyno)       
    query = f"select UNIQUE MASTER_ID from wma.contt where trim(master_id) in "
    if len(policyno) > 0:
        if len(policyno) == 1:   
            query_ = f"{query} (\'{policyno[0]}\')"
            policy_list = runQuery(dbconn,query_)
        else:
            query_ = f"{query} {policyno}"
            policy_list = runQuery(dbconn,query_)
            

    # policy_list = runQuery(dbconn,query_)  
    # print(policy_list)

    if len(policy_list) > 0:
        policy_db_df = pd.DataFrame(policy_list)
        policy_db_df['MASTER_ID'] = policy_db_df['MASTER_ID'].str.strip()
        contt_policies = tuple(policy_db_df['MASTER_ID'])
    else:
        contt_policies = ()

    logger.info('Separating Policies..')  
    conte_policies = tuple(set(policyno) ^ set(contt_policies))
        
    logger.info(f'Policies belongs from contt table are: {contt_policies}')
    logger.info(f'Policies belongs from conte table are: {conte_policies}')


    conte_df = pd.DataFrame()
    contt_df = pd.DataFrame()
    
    # conte_productcodequery = f"SELECT UNIQUE case when a.USER_AREA_1 like '%MPST%' and b.product_code='U1' then 'UL Milepost' when a.USER_AREA_1 like '%MPST%' and b.product_code='U2' then 'VUL Milepost' when a.USER_AREA_1 like '%MPST%' and b.product_code='U3' then 'VOL Milepost' when a.USER_AREA_1 like '%SL    LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%SL    LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%JT01  LCAD%' then 'Joint LifeCadDOS' when a.USER_AREA_1 like '%JT02  LCAD%' then 'Joint LifeCadDOS' when a.USER_AREA_1 like '%SL01  LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%SL02  LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%JT01  SPVL%' then 'Jointlife SPVL' when a.USER_AREA_1 like '%SL01  SPVL%' then 'Singlelife SPVL' when a.USER_AREA_1 like '%SL01  SPD%' then 'Singlelife SPD' when a.USER_AREA_1 like '%JT    LCAD%' then 'Joint LifeCadDOS' when a.USER_AREA_1 like '%SL    SPD%' then 'Singlelife SPD' when a.USER_AREA_1 like '%JT    SPD%' then 'Jointlife SPD' when a.USER_AREA_1 like '%TFTD    SPD%'  then 'TF SPD' when a.USER_AREA_1 like '%SL    SPVL%'   then 'Singlelife SPVL' when a.USER_AREA_1 like '%JT    SPVL%'    then 'Jointlife SPVL' when a.USER_AREA_1 like '%ULS%' then 'ULS' else 'Not rel8' end as product, a.master_id from wma.userx a,wma.conte b where trim(a.master_id) in"
    # conte_Companycodequery = f"select UNIQUE company_code, master_id from wma.conte where trim(master_id) in"
    conte_productcodequery = config['QUERIES']['conte_productcodequery']
    conte_Companycodequery = config['QUERIES']['conte_Companycodequery']
    
    if len(conte_policies) > 0:
        if len(conte_policies) == 1:
            conte_policies = conte_policies[0]
            conte_policies = str(conte_policies)
            
            conte_productcodequery = f"{conte_productcodequery} (\'{conte_policies[0]}\') and trim(b.master_id) in (\'{conte_policies[0]}\') and a.Parent_id = 'DB'"
            conte_Companycodequery =f"{conte_Companycodequery} (\'{conte_policies[0]}\')"

            conte_productcode = runQuery(dbconn,conte_productcodequery)
            conte_Companycode = runQuery(dbconn,conte_Companycodequery)

            
        else:
            conte_productcodequery = f"{conte_productcodequery} {conte_policies} and trim(b.master_id) in {conte_policies} and a.Parent_id = 'DB'"
            conte_Companycodequery =f"{conte_Companycodequery} {conte_policies}"
            
            conte_productcode = runQuery(dbconn,conte_productcodequery)
            conte_Companycode = runQuery(dbconn,conte_Companycodequery)

        conte_productCode_df = pd.DataFrame(conte_productcode)
        conte_companyCode_df = pd.DataFrame(conte_Companycode)

        if conte_productCode_df.empty == False and conte_companyCode_df.empty == False:
            conte_df = pd.merge(conte_productCode_df,conte_companyCode_df, on = 'MASTER_ID')
        else:
            pass        
    else:
        pass
   

    # contt_product_code = "select UNIQUE case when (substr(CONTRACT_USER_AREA, 19, 4) = 'MPST') then 'Milepost' when plan_code like 'I%' Then 'Industrial' when plan_code like 'V%' Then 'VENT/VTRD' when plan_code like 'Q%' Then 'PMF1A' when plan_code in ('P1200','P12001','P12002','P12004','P1205','P12051') Then 'PMF1' when plan_code in ('P1655','P1855','P1023','P10231','P10232','P1572','P15721','P1750','P17501','P17502','P1003','P1008','P1020','P1004','P1014','P1018','P1025','P1075','P1965','P1041','P1105','P1208','P1100','P1150','P1155','P1207','P1050','P1202','P1203','P1250','P1300','P1305','P1036','P1038','P1007','P1043','P1048','P1324','P1321','P1322','P1522','P1529','P1571','P15711','P15712','P1573','P1011','P1016','P1585','P1589','P1590','P1592','P1593','P1594','P1681','P16811','P16812','P1683','P1682','P1684','P1721','P1725','P1771','P1772','P1773','P1785','P1789','P1031','P1086','P1742','P1743','P1745','P1746','P1748','P1749','P1204','P1206','P1209','P0902','P0937','P1055','P1650','P16501','P16502','P16503','P16504','P16505','P16506','P16507','P1674','P1872','P1873','P1675','P1676','P1862','P1868','P1678','P1679','P1671','P1869','P1966','P1690','P1689','P1691','P1734','P1737','P1537','P1534','P1092','P1879','P1850','P18501','P18502','P1657','P1658','P1010','P1074','P1550','P1600','P1625','P1605','P1653','P1605','P1852','P18521','P18522','P1858','P0200','P0209','P0003','P1157','P11571','P11572','P11573','P11574','P1169','P11691','P11692','P11693','P11694','P1306','P13061','P13062','P13063','P13064','P1307','P13071','P13072','P13073','P13074','P1495','P1954','P1955','P1203','P1308','P4044','P4045','P4046','P4047','P4048','P4049','P4083','P4086','P1000','P10001','P1005','P1022','P1027','P1033','P1040','P1045','P1052','P1053','P1057','P1058','P1170','P11701','P1175','P11751','P1180','P11801','P1181','P11811','P1182','P11821','P1183','P11831','P1184','P11841','P1187','P11871','P1188','P11881','P1191','P11911','P1482','P1483','P1484','P1523','P15231','P1524','P15241','P1525','P15251','P1535','P1536','P1574','P15741','P1575','P15751', 'P1576','P1577','P15771','P1578','P15781','P1579','P1660','P16601','P1661','P16611','P1662','P1663','P16631','P1664','P16641','P1665','P1801','P1802','P1803','P1804','P1805','P1806','P1807','P1808','P1809','P1810','P1811','P1812','P4081','P4082','P4084','P4085','P4751','P4752','P4753','P4754','P4755','P4756','P4757','P4758','P4759','P4760','P4761','P4762','P4763','P4764','P4765','P4766','P4767','P4768','P4775','P4776','P4777','P4778','P4779','P4780','P4781','P4782','P4783','P4784','P4785','P4786','P4794','P4795','P4796','P4797','P4798','P4799') Then 'PMF3' else 'PMF2' end product, master_id from wma.contt where trim(master_id) in "
    # contt_company_code = "select UNIQUE company_code, master_id from wma.contt where trim(master_id) in"
    contt_product_code = config['QUERIES']['contt_product_code']
    contt_company_code = config['QUERIES']['contt_company_code']

    if len(contt_policies) > 0:        
        if len(contt_policies) == 1:           
            
            contt_productcodequery = f"{contt_product_code} (\'{contt_policies[0]}\')"
            contt_Companycodequery = f"{contt_company_code} (\'{contt_policies[0]}\')"
            
            contt_productcode = runQuery(dbconn,contt_productcodequery)
            contt_Companycode = runQuery(dbconn,contt_Companycodequery)

        else:
            contt_productcodequery = f"{contt_product_code} {contt_policies}"
            contt_Companycodequery = f"{contt_company_code} {contt_policies}"
            
            contt_productcode = runQuery(dbconn,contt_productcodequery)
            contt_Companycode = runQuery(dbconn,contt_Companycodequery)

        contt_productCode_df = pd.DataFrame(contt_productcode)
        contt_companyCode_df = pd.DataFrame(contt_Companycode)

        if contt_productCode_df.empty == False and contt_companyCode_df.empty == False:
            contt_df = pd.merge(contt_productCode_df,contt_companyCode_df, on = 'MASTER_ID')     
        else:
            pass   
        
    else:
        pass


    if contt_df.empty == False and conte_df.empty == False:        
        final_df = contt_df.append(conte_df, ignore_index=True)
        final_df = final_df.rename(columns={'MASTER_ID':'Policy_No','PRODUCT':'Product_Type','COMPANY_CODE':'Company_Code'})

    elif contt_df.empty == False and conte_df.empty == True:
        final_df = contt_df
        final_df = final_df.rename(columns={'MASTER_ID':'Policy_No','PRODUCT':'Product_Type','COMPANY_CODE':'Company_Code'})
    elif contt_df.empty == True and conte_df.empty == False:
        final_df = conte_df
        final_df = final_df.rename(columns={'MASTER_ID':'Policy_No','PRODUCT':'Product_Type','COMPANY_CODE':'Company_Code'})
        
    else:
        print('Policies are not found in the database')
        logger.info('Policies are not found in the database')
        final_df = pd.DataFrame()

    # print(final_df)
    # final_df.to_excel('finalDF.xlsx',index = False)
    return final_df   

def createExcel(df1,df2,availablePDF):        
    file_time = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
    filename = f'{resultpath}Documaker Policy Details_{file_time}'
    with pd.ExcelWriter(f'{filename}.xlsx') as writer:  
        df1.to_excel(writer, index = False, sheet_name='DocumakerCount')
        df2.to_excel(writer, index = False, sheet_name='DuplicatePolicies')
        availablePDF.to_excel(writer, index = False, sheet_name = 'PDFs')

    
    #Format the columns of excel file 
    #Auto adjust column width       
    wb = openpyxl.load_workbook(filename = f'{filename}.xlsx')
    for sheet in wb.sheetnames:       
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name        
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width         
                                                        
        wb.save(f'{filename}.xlsx')


def run_me():
    Startpagevalue = config['PAGERANGE']['startpage']
    Endrangevalue = config['PAGERANGE']['endpage']

    # GET THE TOTAL NUMBER OF PDF FROM downloadpdf FOLDER
    available_pdf = fnmatch.filter(os.listdir(pdflocation), '*.pdf')     
    
    # PACKING ALL THE VALUES TO PASS WITH CONCURRENT EXECUTOR
    all_params = [(pdf,Startpagevalue,Endrangevalue) for pdf in available_pdf]

    print('STARTING PARALLEL EXECUTION...')
    logger.info('STARTING PARALLEL EXECUTION...')

    # STARTING CONCURRENT PROCESS AND PASSING THE VALUES NEEDED FOR FURTHER EXECUTION
    with concurrent.futures.ProcessPoolExecutor(max_workers=len(available_pdf)) as executor:
        content = executor.map(parallel_execution, all_params)       
    
    print('PARALLEL EXECUTION ENDED...')
    logger.info('PARALLEL EXECUTION ENDED...')
    data_list = []

    for c in content:
        data_list.extend(c)

    df = pd.DataFrame(data_list)   
    # print(df) 
    df.to_excel('a.xlsx', index = False)
    
    df1 = df.drop_duplicates(subset=['Policy_No'],ignore_index=True)
    df1.to_excel('b.xlsx', index = False)
    # print(df1)
    logger.info('Data Frame Extracted.')
    new_df = df.loc[df.duplicated()].copy()
    # print(new_df)
    new_df.to_excel('c.xlsx', index = False)
    logger.info('Duplicate Data Separated.')
    
    db_strattime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
    # STABLISH DB CONNECTION
    try:
        print(f'Connecting IBM DB2: {hostName} - Initiated')
        dbIBMConn = ibm_db.connect(f'DATABASE={dbName};HOSTNAME={hostName};PORT={port};PROTOCOL=TCPIP;UID={uDname};PWD={passWord};', "", "")
        
    except Exception as e:
        print(f'ERROR dbConnIBM - CONNECTING IBM DB\n--> Please Check the error messgae<-- Error Message: {e}')        
        logger.error(f'ERROR dbConnIBM CONNECTING IBM DB2 SERVER: {hostName}, DBName: {dbName}, UID={uDname}, PWD={passWord}\n{e}')
    
    else:
        print(f'IBM DB2 Connection - Success')
        logger.info(f'IBM DB2 Connection to {hostName} - Success')
        
        print('Please Wait... Fetching Details from Data base.')
        # df1.to_excel('b.xlsx', index = False)
        df1['Policy_No'] = df1['Policy_No'].str.strip()
        policy_col = tuple(df1['Policy_No'])

        
        logger.info(f'Extracted the Policy Numbers: {policy_col}')       
       
        final_dataframe = handle_db_query(dbIBMConn,policy_col) 

        db_endtime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
        print(f"\nDB_Start Time: {db_strattime}, DB_End Time: {db_endtime}\n")         
  
        # df1['Policy_No'] = df1['Policy_No'].str.strip()
        final_dataframe['Policy_No'] = final_dataframe['Policy_No'].str.strip()
        
        merged_df = pd.merge(df1,final_dataframe, on = 'Policy_No') 
        # print(f'merged_df:\n{merged_df}')

        merged_df = merged_df[['Letter_Name','Letter_Heading', 'Policy_No', 'Company_Code','Product_Type','Pdf_Name']]
        pdf_df = pd.DataFrame(available_pdf,columns=['Available_PDF'])
        merged_df = merged_df.sort_values(by = 'Letter_Name')
        new_df = new_df.sort_values(by = 'Letter_Name')
        pdf_df = pdf_df.sort_values(by = 'Available_PDF')
        createExcel(merged_df,new_df,pdf_df)
        print('Excel Created. Please check the Result Folder')
        logging.info('Excel Created. Please check the Result Folder')


if __name__ == '__main__':   
    freeze_support()    
    strattime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")

    move_file()
    run_me()

    endtime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
    logger.info(f'Total time taken Start time: {strattime} and Endtime {endtime}')
    print(f"\nStart Time: {strattime}, End Time: {endtime}")



    # newdf = df[df.duplicated(['Policy_No','Letter_Name','Pdf_Name'], keep = 'first')]
    
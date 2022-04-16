from multiprocessing import allow_connection_pickling, freeze_support
import configparser, logging
from logging.handlers import TimedRotatingFileHandler
import shutil, os,os.path,fnmatch, re, sys
from io import StringIO
from typing_extensions import ParamSpec
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
import pandas as pd
import concurrent
from concurrent.futures import ProcessPoolExecutor
import ibm_db
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import date, datetime
import warnings,json

warnings.filterwarnings("ignore")

# SETTUP FOR LOGGING
logger = logging.getLogger('DocumakerCount')    
logger.setLevel(logging.INFO)
rotate = TimedRotatingFileHandler(f'LOGS\\DocumakerCount.log', when='D', 
                                    interval=1, backupCount=20, encoding=None, 
                                    delay=False, utc=False)

logger.addHandler(rotate)
formater = logging.Formatter('%(asctime)s:%(levelname)s:%(name)s:%(message)s')
rotate.setFormatter(formater)

# SETUP CONFIG PARSER TO READ VALUES FROM INI FILE
config = configparser.RawConfigParser()
config.read('countproperties.ini')

pdflocation = f'downloadpdf\\'
resultpath = f'Results\\'

hostName = config['DBCONNECT']['host']
port = config['DBCONNECT']['port']
dbName = config['DBCONNECT']['dbname']
uDname = config['DBCONNECT']['dbuser']
passWord = config['DBCONNECT']['dbpass']

# MOVE OLD FILE FROM RESULT FOLDER TO BACKUP FOLDER
def move_file():
    backup = f'Results\Backup'
    file = fnmatch.filter(os.listdir(resultpath), '*.xlsx')
    if len(file) > 0:
        shutil.move(f'{resultpath}{file[0]}',backup)
        logger.info('Line No: 48: Program Started. Excel File moved to Backup Folder.')

def read_pdf(filename,Startpagevalue,Endrangevalue):      
    output_string01 = StringIO()      
    
    with open(f'{filename}', 'rb') as in_file:
        in_file = open(f'{filename}', 'rb')
        parser = PDFParser(in_file)        
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()        
        device = TextConverter(rsrcmgr, output_string01, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        
        
        for i, page in enumerate(PDFPage.create_pages(doc)):
            if Startpagevalue=='0' and Endrangevalue == '0':
                read_position = output_string01.tell()
                interpreter.process_page(page)
                output_string01.seek(read_position, 0)
                page_text = output_string01.read()
                eachpagelist = re.split('\n+', page_text)
                eachpagelist = [x for x in eachpagelist if int(len(x)) > 1]                
                yield eachpagelist
            else:
                if i>=int(Startpagevalue) and i<=int(Endrangevalue):
                    read_position = output_string01.tell()
                    interpreter.process_page(page)
                    output_string01.seek(read_position, 0)
                    page_text = output_string01.read()
                    eachpagelist = re.split('\n+', page_text)
                    eachpagelist = [x for x in eachpagelist if int(len(x)) > 1]                    
                    yield eachpagelist
                else:
                    pass

def extractAllDetails(key,filename,letterscount,letterheading,lowerlistToStrvalue,max_no_of_alphabets_in_policy):
    lettername = key
    if(lettername in("stale check")):
        if filename.__contains__("STALECHECK90"):
            lettername = 'stalecheck90'
        else:
            lettername = 'stalecheck150'
        # if lettername in set1:
        #     firstvalue = "policy number"                   


        # elif lettername in set2:
        #     if lowerlistToStrvalue.__contains__('policy:'):
        #         firstvalue = "policy:"
        #     else:
        #         firstvalue = "company policy"                    
        
        # else:
        #     firstvalue = 'policy number'    
    letterscount = letterscount+1
    
    row_dict = {}
    try:
        policynumber = ""
        firstvalue = ('policy number' if ('policy number' in lowerlistToStrvalue.lower()) else
                    'policy:' if ('policy:' in lowerlistToStrvalue.lower()) else
                    'company policy' if ('company policy' in lowerlistToStrvalue.lower()) else
                    'policy  ' if ('policy  ' in lowerlistToStrvalue.lower()) else
                    'policy '
                    ).strip()

        fv = "_".join(firstvalue.split(' ')) if (len(firstvalue.split(' ')) > 1) else firstvalue
        lowerlistToStrvalue = lowerlistToStrvalue.lower().replace(firstvalue, fv)
        lowerlist = lowerlistToStrvalue.split(' ')
        # print(lowerlist)
        findindex = lowerlist.index(fv)
        # print(fv)
        indices = []
        for i in range(len(lowerlist)):
            if 'policy' in lowerlist[i]:
                indices.append(i)
        for findindex in indices:
            var = lowerlist[findindex+1]
            if len(var) >= 6 and len(var) <= 15:
                rx = re.compile('\w{6,15}')
                exp = rx.match(var).group()
                if exp:
                    maxval = len(exp)
                    digCount = re.findall('\d', exp)
                    alpCount = re.findall('\D', exp)
                    max_no_of_alphabets_in_policy = 3
                    if len(digCount) != 0 or (len(alpCount) >= 0 and len(alpCount) <= int(max_no_of_alphabets_in_policy)):                
                        policynumber = (str(exp)).upper()
                        print(policynumber)
                        break               
                        
        logger.info(f'Letter_Name: {lettername}, Policy_Number: {policynumber}')

        row_dict = {
            'Letter_Name' : lettername,
            'Letter_Heading' : letterheading,
            'Policy_No' : policynumber,
            'Pdf_Name' : filename
        }
                        
                        
    except Exception as e:
        logger.error(f'Error while processing File_Name:{filename}, Letter_Name: {key} and Policy_Number: {policynumber}, because of:  {e}')
            
            
    return row_dict


def extract_contents(filename,Startpagevalue,Endrangevalue,letternameandvalue,max_no_of_alphabets_in_policy):
    path = f'downloadpdf\\'
    print(f'Processing PDF: {filename}')
    # print(path)
    datas = list(read_pdf(f'{path}{filename}',Startpagevalue,Endrangevalue))
    # print(len(datas
    # print(datas)
        

    # letternameandvalue = {'notice third party offer':['important information about your policy - please review','notice third party offer'],'confirmation premium payment':['premium payment confirmation'],
    # 'confirmation loan collateral interest credit':['loan collateral interest credit confirmation'],'confirmation loan interest capitalization':['loan interest capitalization confirmation'],'bill loan interest only':['notice of loan interest payment due','face amount'],'confirmation apl':['automatic premium loan confirmation'],'notice net gain dividends':['important tax information – immediate attention needed'],'confirmation discontinue accelerated payment-premium offset':["discontinue accelerated payment confirmation"],'notice term expiration':['notice term expiration'],'quote withdrawl':['policy withdrawal quote'],'audit':['your request has been received – please review','audit'],'1035 exchange confirmation':['1035 exchange confirmation'],'acknowledgement of life insurance letter':['statement of coverage – please review'],'accelerated payment confirmation':['accelerated payment confirmation','under this payment arrangement, a withdrawal from your policy’s'],'accelerated payment quote':['accelerated payment quote'],'address change confirmation':['address change confirmation'],'annual policy statement':['annual policy statement'],'auto reinstatement':['notice of policy lapse','thank you for your premium payment'],'bill grace prelapse':['notice of overdue payment','refer to your original policy for specific'],
    # 'ca beneficiary notification':['ca beneficiary notification','important information about your policy– please review'],'ca consumer letter':['important insurance disclosure – please review','ca consumer letter'],'confirmation frequency change':['payment change confirmation','the frequency of'],'confirmation reinstatement':['reinstatement confirmation'],'dividend option change':['dividend transaction confirmation','dividend transfer confirmation','dividend option change confirmation'],'electronic payment confirmation':['electronic payment confirmation'],'explanation of benefits':['explanation of benefits','total control account (tca)'],'fl consumer letter':['fl consumer letter','important insurance disclosure – please review'],'fully paid-up confirmation':['fully paid-up confirmation'],'life quote nonforfeiture':['reduced paid-up insurance quote','life quote nonforfeiture'],'loan exceeds cash value termination':['important information – your policy has been cancelled'],'lapse notification':['important notification – your policy has lapsed'],'loan interest change confirmation':['loan interest change confirmation'],'loan repayment confirmation':['loan repayment confirmation'],'maturity notification':['your reply is important – please review and respond'],'nh consumer letter':['nh consumer letter','Important insurance disclosure – Please review'],'notice impact of loan or withdrawal on accelerated payment':['important information – please review'],'notice child':['important information about your policy– please review','this notice is to remind you that you have a child term rider'],'notice nsf':['your payment has been returned– immediate attention required'],'notice of minimum loan repayment':['notice of minimum loan repayment'],'owner and beneficiary change':['policy update confirmation','collateral assignment confirmation','name change confirmation','owner confirmation','beneficiary change confirmation'],'policy cash value quote':['policy cash value quote'],'policy dividend history' :['policy dividend history'],'policy loan confirmation':['policy loan confirmation'],'policy loan history':['policy loan history'],
    # 'policy loan quote':['policy loan quote'],'policy surrender confirmation':['policy surrender confirmation'],'premium payment history':['premium payment history'],'quote term policy':['quote term policy','term policy value quote'],'ri consumer letter':['ri consumer letter','important insurance disclosure – please review'],'recorded nonforfeiture confirmation':['recorded nonforfeiture confirmation'],'skip recovery privacy notice':['privacy policy – please review'],'stale check':['your request has been received – please review','stale dated checks'],'tax identification confirmation':['tax identification number update confirmation'],'third party designation confirmation':['third party designation confirmation'],'wi consumer letter':['important insurance disclosure – please review','wi consumer letter'],'wma check':['pay to the order of','reason:'],'expiration of rider-benefit':['your policy contains a benefit – please review','your policy contains a rider – please review'],
    # 'notice premium or rider premium change':['your premium is changing – please review','what you need to do'],'suspend billing confirmation':['suspend billing confirmation','resume billing confirmation'],'notice of payment due':['notice of payment due','face amount'],'confirmation planned premium change':['payment change confirmation','we processed your premium change'],'confirmation annual dividend election':['annual dividend election confirmation'],'confirmation fund transfer':['fund transfer confirmation'],'confirmation rebalancing program':['rebalancing program confirmation'],'confirmation premium allocation':['premium allocation confirmation'],'confirmation loan interest rate change':['loan interest change confirmation'],'retail life annual statement':['statement_name']}

    letterscount = 0
    all_data = []
    for data in datas:
        listToStrvalue = ' '.join([str(elem) for elem in data])
        
        # Remove unwanted spaces
        listToStrvalue = re.sub('\s{2,}', ' ', listToStrvalue)
        # print(f'{listToStrvalue}')
        data_dict = {}
        logger.info(listToStrvalue)
        for key in letternameandvalue:
            letterheadinglist = letternameandvalue[key]
            lowerlistToStrvalue = listToStrvalue.lower()            
            # print(lowerlistToStrvalue)
            # sys.exit()            
            
            
            present = all(item in lowerlistToStrvalue for item in letterheadinglist)           

            try:
                # Covered the NOT Condition
                if key == 'confirmation frequency change':
                    if present == True:
                        if 'this is not a bill' not in lowerlistToStrvalue:
                            letterheading = letterheadinglist[0]
                            data_dict = extractAllDetails(key,filename,letterscount,letterheading,lowerlistToStrvalue,max_no_of_alphabets_in_policy)
                            all_data.append(data_dict)

                # Covered the NOT Condition
                if key == 'auto reinstatement':
                    if present == True:                   
                        for item in letterheadinglist:
                            if (item in lowerlistToStrvalue) and ('your representative' not in lowerlistToStrvalue):
                                letterheading = item
                    # if 'Your Representative' not in lowerlistToStrvalue:
                                data_dict = extractAllDetails(key,filename,letterscount,letterheading,lowerlistToStrvalue,max_no_of_alphabets_in_policy)
                                all_data.append(data_dict)
                
                # Covered the NOT Condition
                if key == 'notice of payment due- loan reminder':
                    if present == True:
                        if 'important information about your loan repayment reminder' not in lowerlistToStrvalue:
                            letterheading = letterheadinglist[0]
                            data_dict = extractAllDetails(key,filename,letterscount,letterheading,lowerlistToStrvalue,max_no_of_alphabets_in_policy)
                            all_data.append(data_dict)


                # OR Condition
                if present == False:
                    if key in ['dividend option change','expiration of rider-benefit','owner and beneficiary change','notice term expiration']:                   
                        for item in letterheadinglist:
                            if item in lowerlistToStrvalue:
                                letterheading = item
                                data_dict = extractAllDetails(key,filename,letterscount,letterheading,lowerlistToStrvalue,max_no_of_alphabets_in_policy)
                                all_data.append(data_dict)
                            else:
                                pass

                # AND Condition 
                if present == True:
                    letterheading = letterheadinglist[0]
                    if key != 'confirmation frequency change' and key != 'auto reinstatement' and key!= 'notice of payment due- loan reminder':
                        data_dict = extractAllDetails(key,filename,letterscount,letterheading,lowerlistToStrvalue,max_no_of_alphabets_in_policy)
                        all_data.append(data_dict)
                    else:
                        pass                

            except Exception as e:
                logger.error(f'Error while processing File_Name:{filename}, Letter_Name: {key} because of:  {e}')
                # logger.error(f'Error while processing File_Name:{filename}, Letter_Name: {key} and Policy_Number: {data_dict["Policy_No"]} because of:  {e}')
                # logger.info(data)
                # sys.exit()
                continue
            # finally:
            #     if present == True:                    
            #         sys.exit()    
    return all_data


def parallel_execution(params_list):    
    required_content = extract_contents(params_list[0],params_list[1],params_list[2],params_list[3],params_list[4])    
    return required_content


# METHOD TO RUN DB QUERY 
def runQuery(dbIBMConn,query):
    try:        
        stmtprp = ibm_db.exec_immediate(dbIBMConn, query)        
        rowData = ibm_db.fetch_assoc(stmtprp)       
 
    except Exception as e:
        print(f'ERROR runQueryIBM - {e}')
        logger.error(f'ERROR runQueryIBM - {e}')
        
    else:     
        logger.info('Query ran sucessfully. Tabulating the data...')     
        ibmResultList = []
        while rowData != False:
            ibmResultList.append(rowData)
            rowData = ibm_db.fetch_assoc(stmtprp)   
        # print(ibmResultList)   
        # logger.info(ibmResultList)      
        return ibmResultList


def handle_db_query(dbconn,policyno): 
    # print(policyno)       
    query = f"select UNIQUE MASTER_ID from wma.contt where trim(master_id) in "
    if len(policyno) > 0:
        if len(policyno) == 1:   
            query_ = f"{query} (\'{policyno[0]}\')"
            policy_list = runQuery(dbconn,query_)
            
        else:            
            query_ = f"{query} {policyno}"
            policy_list = runQuery(dbconn,query_)
            

    # policy_list = runQuery(dbconn,query_)  
    # print(policy_list)

    if len(policy_list) > 0:
        policy_db_df = pd.DataFrame(policy_list)
        policy_db_df['MASTER_ID'] = policy_db_df['MASTER_ID'].str.strip()
        contt_policies = tuple(policy_db_df['MASTER_ID'])
    else:
        contt_policies = ()

    logger.info('Separating Policies..')  
    conte_policies = tuple(set(policyno) ^ set(contt_policies))
        
    logger.info(f'Policies belongs from contt table are: {contt_policies}')
    logger.info(f'Policies belongs from conte table are: {conte_policies}')


    conte_df = pd.DataFrame()
    contt_df = pd.DataFrame()
    
    # conte_productcodequery = f"SELECT UNIQUE case when a.USER_AREA_1 like '%MPST%' and b.product_code='U1' then 'UL Milepost' when a.USER_AREA_1 like '%MPST%' and b.product_code='U2' then 'VUL Milepost' when a.USER_AREA_1 like '%MPST%' and b.product_code='U3' then 'VOL Milepost' when a.USER_AREA_1 like '%SL    LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%SL    LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%JT01  LCAD%' then 'Joint LifeCadDOS' when a.USER_AREA_1 like '%JT02  LCAD%' then 'Joint LifeCadDOS' when a.USER_AREA_1 like '%SL01  LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%SL02  LCAD%' then 'Single LifeCadDOS' when a.USER_AREA_1 like '%JT01  SPVL%' then 'Jointlife SPVL' when a.USER_AREA_1 like '%SL01  SPVL%' then 'Singlelife SPVL' when a.USER_AREA_1 like '%SL01  SPD%' then 'Singlelife SPD' when a.USER_AREA_1 like '%JT    LCAD%' then 'Joint LifeCadDOS' when a.USER_AREA_1 like '%SL    SPD%' then 'Singlelife SPD' when a.USER_AREA_1 like '%JT    SPD%' then 'Jointlife SPD' when a.USER_AREA_1 like '%TFTD    SPD%'  then 'TF SPD' when a.USER_AREA_1 like '%SL    SPVL%'   then 'Singlelife SPVL' when a.USER_AREA_1 like '%JT    SPVL%'    then 'Jointlife SPVL' when a.USER_AREA_1 like '%ULS%' then 'ULS' else 'Not rel8' end as product, a.master_id from wma.userx a,wma.conte b where trim(a.master_id) in"
    # conte_Companycodequery = f"select UNIQUE company_code, master_id from wma.conte where trim(master_id) in"
    conte_productcodequery = config['QUERIES']['conte_productcodequery']
    conte_Companycodequery = config['QUERIES']['conte_Companycodequery']
    
    if len(conte_policies) > 0:
        if len(conte_policies) == 1:
            conte_policies = conte_policies[0]
            conte_policies = str(conte_policies)
            # print(conte_policies[0])
            
            
            conte_productcodequery = f"{conte_productcodequery} (\'{conte_policies}\') and trim(b.master_id) in (\'{conte_policies}\') and a.Parent_id = 'DB'"
            conte_Companycodequery = f"{conte_Companycodequery} (\'{conte_policies}\')"
              
            conte_productcode = runQuery(dbconn,conte_productcodequery)
            conte_Companycode = runQuery(dbconn,conte_Companycodequery)
            
            
        else:
            conte_productcodequery = f"{conte_productcodequery} {conte_policies} and trim(b.master_id) in {conte_policies} and a.Parent_id = 'DB'"
            conte_Companycodequery = f"{conte_Companycodequery} {conte_policies}"
            
            conte_productcode = runQuery(dbconn,conte_productcodequery)
            conte_Companycode = runQuery(dbconn,conte_Companycodequery)
            # print(conte_Companycode)
            

        conte_productCode_df = pd.DataFrame(conte_productcode)
        conte_companyCode_df = pd.DataFrame(conte_Companycode)               

        if conte_productCode_df.empty == False and conte_companyCode_df.empty == False:
            conte_df = pd.merge(conte_productCode_df,conte_companyCode_df, on = 'MASTER_ID')
            # print(f'conte_df: {conte_df}')
            
        else:
            pass        
    else:
        pass
   

    # contt_product_code = "select UNIQUE case when (substr(CONTRACT_USER_AREA, 19, 4) = 'MPST') then 'Milepost' when plan_code like 'I%' Then 'Industrial' when plan_code like 'V%' Then 'VENT/VTRD' when plan_code like 'Q%' Then 'PMF1A' when plan_code in ('P1200','P12001','P12002','P12004','P1205','P12051') Then 'PMF1' when plan_code in ('P1655','P1855','P1023','P10231','P10232','P1572','P15721','P1750','P17501','P17502','P1003','P1008','P1020','P1004','P1014','P1018','P1025','P1075','P1965','P1041','P1105','P1208','P1100','P1150','P1155','P1207','P1050','P1202','P1203','P1250','P1300','P1305','P1036','P1038','P1007','P1043','P1048','P1324','P1321','P1322','P1522','P1529','P1571','P15711','P15712','P1573','P1011','P1016','P1585','P1589','P1590','P1592','P1593','P1594','P1681','P16811','P16812','P1683','P1682','P1684','P1721','P1725','P1771','P1772','P1773','P1785','P1789','P1031','P1086','P1742','P1743','P1745','P1746','P1748','P1749','P1204','P1206','P1209','P0902','P0937','P1055','P1650','P16501','P16502','P16503','P16504','P16505','P16506','P16507','P1674','P1872','P1873','P1675','P1676','P1862','P1868','P1678','P1679','P1671','P1869','P1966','P1690','P1689','P1691','P1734','P1737','P1537','P1534','P1092','P1879','P1850','P18501','P18502','P1657','P1658','P1010','P1074','P1550','P1600','P1625','P1605','P1653','P1605','P1852','P18521','P18522','P1858','P0200','P0209','P0003','P1157','P11571','P11572','P11573','P11574','P1169','P11691','P11692','P11693','P11694','P1306','P13061','P13062','P13063','P13064','P1307','P13071','P13072','P13073','P13074','P1495','P1954','P1955','P1203','P1308','P4044','P4045','P4046','P4047','P4048','P4049','P4083','P4086','P1000','P10001','P1005','P1022','P1027','P1033','P1040','P1045','P1052','P1053','P1057','P1058','P1170','P11701','P1175','P11751','P1180','P11801','P1181','P11811','P1182','P11821','P1183','P11831','P1184','P11841','P1187','P11871','P1188','P11881','P1191','P11911','P1482','P1483','P1484','P1523','P15231','P1524','P15241','P1525','P15251','P1535','P1536','P1574','P15741','P1575','P15751', 'P1576','P1577','P15771','P1578','P15781','P1579','P1660','P16601','P1661','P16611','P1662','P1663','P16631','P1664','P16641','P1665','P1801','P1802','P1803','P1804','P1805','P1806','P1807','P1808','P1809','P1810','P1811','P1812','P4081','P4082','P4084','P4085','P4751','P4752','P4753','P4754','P4755','P4756','P4757','P4758','P4759','P4760','P4761','P4762','P4763','P4764','P4765','P4766','P4767','P4768','P4775','P4776','P4777','P4778','P4779','P4780','P4781','P4782','P4783','P4784','P4785','P4786','P4794','P4795','P4796','P4797','P4798','P4799') Then 'PMF3' else 'PMF2' end product, master_id from wma.contt where trim(master_id) in "
    # contt_company_code = "select UNIQUE company_code, master_id from wma.contt where trim(master_id) in"
    contt_product_code = config['QUERIES']['contt_product_code']
    contt_company_code = config['QUERIES']['contt_company_code']
    
    if len(contt_policies) > 0:        
        if len(contt_policies) == 1: 
            contt_policies = contt_policies[0]
            contt_policies = str(contt_policies)          
            
            contt_productcodequery = f"{contt_product_code} (\'{contt_policies}\')"
            contt_Companycodequery = f"{contt_company_code} (\'{contt_policies}\')"                       
            
            contt_productcode = runQuery(dbconn,contt_productcodequery)
            contt_Companycode = runQuery(dbconn,contt_Companycodequery)

        else:
            contt_productcodequery = f"{contt_product_code} {contt_policies}"
            contt_Companycodequery = f"{contt_company_code} {contt_policies}"
            
            contt_productcode = runQuery(dbconn,contt_productcodequery)
            contt_Companycode = runQuery(dbconn,contt_Companycodequery)

        contt_productCode_df = pd.DataFrame(contt_productcode)
        contt_companyCode_df = pd.DataFrame(contt_Companycode)

        if contt_productCode_df.empty == False and contt_companyCode_df.empty == False:
            contt_df = pd.merge(contt_productCode_df,contt_companyCode_df, on = 'MASTER_ID')     
        else:
            pass   
        
    else:
        pass


    if contt_df.empty == False and conte_df.empty == False:        
        final_df = contt_df.append(conte_df, ignore_index=True)
        final_df = final_df.rename(columns={'MASTER_ID':'Policy_No','PRODUCT':'Product_Type','COMPANY_CODE':'Company_Code'})
        # print(final_df)
    elif contt_df.empty == False and conte_df.empty == True:
        final_df = contt_df
        final_df = final_df.rename(columns={'MASTER_ID':'Policy_No','PRODUCT':'Product_Type','COMPANY_CODE':'Company_Code'})
    elif contt_df.empty == True and conte_df.empty == False:
        final_df = conte_df
        final_df = final_df.rename(columns={'MASTER_ID':'Policy_No','PRODUCT':'Product_Type','COMPANY_CODE':'Company_Code'})
        
    else:
        print('Policies are not found in the database')
        logger.info('Policies are not found in the database')
        column_names = ['Policy_No','Product_Type','Company_Code']
        final_df = pd.DataFrame(columns = column_names)

    return final_df   

# METHOD TO CREATE EXCEL FILE AND ADJUST COLUMMN WIDTH AUTOMATICALLY
def createExcel(df1,df2,availablePDF):        
    file_time = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
    filename = f'{resultpath}Documaker Policy Details_{file_time}'
    with pd.ExcelWriter(f'{filename}.xlsx') as writer:  
        df1.to_excel(writer, index = False, sheet_name='DocumakerCount')
        df2.to_excel(writer, index = False, sheet_name='DuplicatePolicies')
        availablePDF.to_excel(writer, index = False, sheet_name = 'PDFs')

    
    #Format the columns of excel file 
    #Auto adjust column width       
    wb = openpyxl.load_workbook(filename = f'{filename}.xlsx')
    for sheet in wb.sheetnames:       
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name        
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width         
                                                        
        wb.save(f'{filename}.xlsx')

# METHOD FOR ALL THE LOGICAL OPERATIONS
def run_me():
    Startpagevalue = config['PAGERANGE']['startpage']
    Endrangevalue = config['PAGERANGE']['endpage']
    max_no_of_alphabets_in_policy = config['ALPHABETS_COUNT_IN_POILICY']['max_no_of_alphabets_in_policy']


    # GET THE TOTAL NUMBER OF PDF FROM downloadpdf FOLDER
    available_pdf = fnmatch.filter(os.listdir(pdflocation), '*.pdf')

    # OPEN JSON FILE AND CONVERT IT TO DICTIONARY FOR FURTHER EXECUTION, ENCODE IT TO utf-8 BECOZ FILE CONTAINS SOME SPECIAL CHARACTER ALSO
    with open("LetterName.json", encoding='utf-8') as json_file:
        json_data = json.load(json_file)
        json_data =  {k.lower(): [value.lower() for value in v] for k, v in json_data.items()}
        # print(json_data)
        # sys.exit()
    
    # PACKING ALL THE VALUES TO PASS WITH CONCURRENT EXECUTOR
    all_params = [(pdf,Startpagevalue,Endrangevalue,json_data,max_no_of_alphabets_in_policy) for pdf in available_pdf]

    print('STARTING PARALLEL EXECUTION...')
    logger.info('STARTING PARALLEL EXECUTION...')

    # STARTING CONCURRENT PROCESS AND PASSING THE VALUES NEEDED FOR FURTHER EXECUTION
    with concurrent.futures.ProcessPoolExecutor(max_workers=len(available_pdf)) as executor:
        content = executor.map(parallel_execution, all_params)       
    
    print('PARALLEL EXECUTION ENDED...')
    logger.info('PARALLEL EXECUTION ENDED...')
    data_list = []

    for c in content:
        data_list.extend(c)

    
    df = pd.DataFrame(data_list)   

    db_strattime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
    # STABLISH DB CONNECTION
    try:
        print(f'Connecting IBM DB2: {hostName} - Initiated')
        dbIBMConn = ibm_db.connect(f'DATABASE={dbName};HOSTNAME={hostName};PORT={port};PROTOCOL=TCPIP;UID={uDname};PWD={passWord};', "", "")
        
    except Exception as e:
        print(f'ERROR dbConnIBM - CONNECTING IBM DB\n--> Please Check the error messgae<-- Error Message: {e}')        
        logger.error(f'ERROR dbConnIBM CONNECTING IBM DB2 SERVER: {hostName}, DBName: {dbName}, UID={uDname}, PWD={passWord}\n{e}')
    
    else:
        print(f'IBM DB2 Connection - Success')
        logger.info(f'IBM DB2 Connection to {hostName} - Success')        
        print('Please Wait... Fetching Details from Data Base.')
        
        
        # print(df)
        df['Policy_No'] = df['Policy_No'].str.strip()  
        nan_value = float("NaN")
        df.replace("", nan_value, inplace=True)
        df.dropna(subset = ["Policy_No"], inplace=True)      
        policy_col = tuple(df['Policy_No'])
        
        
        logger.info(f'Extracted the Policy Numbers: {policy_col}')       
       
        final_dataframe = handle_db_query(dbIBMConn,policy_col)

        db_endtime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
        print(f"\nDB_Start Time: {db_strattime}, DB_End Time: {db_endtime}\n")        
        

        final_dataframe['Policy_No'] = final_dataframe['Policy_No'].str.strip()        
        codes_list = final_dataframe.to_dict('records')
        
        df['Product_Type'] = ''
        df['Company_Code'] = ''

        for i in df.index: 
            p = (df['Policy_No'][i]).strip()
            res = [[sub['Product_Type'],sub['Company_Code']] for sub in codes_list if sub['Policy_No'] == p]
            # print(res[0])
            if len(res) != 0:
                df.loc[i,'Product_Type'] = res[0][0]
                df.loc[i,'Company_Code'] = res[0][1]
        
        # print(df)
        # newdf = df[df.duplicated(['Policy_No','Letter_Name','Pdf_Name'], keep = 'first')]
        newdf = df[df.duplicated(subset = ['Policy_No','Letter_Name','Pdf_Name'], keep = False)]
        df  = df.drop_duplicates(subset = ['Policy_No','Letter_Name','Pdf_Name'], keep = 'first')
        pdf_df = pd.DataFrame(available_pdf,columns=['Available_PDF'])

        # Remove the 1st dupliacte element from newdf
        newdfdrop = newdf[~newdf.duplicated(keep='first')]
        filtdf = newdf.drop(list(newdfdrop.index))
        # print(filtdf)

        df = df[['Letter_Name','Letter_Heading', 'Policy_No', 'Company_Code','Product_Type','Pdf_Name']]
        filtdf = filtdf[['Letter_Name','Letter_Heading', 'Policy_No', 'Company_Code','Product_Type','Pdf_Name']]

        df = df.sort_values(by = 'Letter_Name')
        filtdf = filtdf.sort_values(by = 'Letter_Name')
        pdf_df = pdf_df.sort_values(by = 'Available_PDF')
        createExcel(df,filtdf,pdf_df)
        
        print('Excel Created. Please check the Result Folder')
        logging.info('Excel Created. Please check the Result Folder')


if __name__ == '__main__':   
    freeze_support()
    logger.info('####################### Execution Started #########################')    
    strattime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")

    move_file()
    run_me()

    endtime = datetime.now().strftime("%Y-%m-%d_%I-%M-%S_%p")
    logger.info(f'Total time taken Start time: {strattime} and Endtime {endtime}')
    print(f"\nStart Time: {strattime}, End Time: {endtime}")
    logger.info('####################### Execution Completed #########################')

